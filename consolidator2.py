import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re
import argparse
import sys

class FinancialStatementConsolidator:
    def __init__(self, statements_dir):
        self.statements_dir = statements_dir
        self.files = {
            'balance_sheet': {
                'FY': None,
                'QTR': None
            },
            'income_statement': {
                'FY': None,
                'QTR': None
            },
            'cash_flow': {
                'FY': None,
                'QTR': None
            }
        }
        
        # Extract ticker from the directory name
        self.company_ticker = os.path.basename(statements_dir).upper()
        self._load_files()

    def _load_files(self):
        """Load all Excel files and categorize them."""
        for filename in os.listdir(self.statements_dir):
            if not filename.endswith('.xlsx') or filename.startswith('~'):
                continue
                
            period_type = filename.split('_')[0]  # FY or QTR
            
            # No need to extract ticker from filename anymore, as we get it from directory
            
            if 'balance_sheet' in filename.lower():
                statement_type = 'balance_sheet'
            elif 'income_statement' in filename.lower():
                statement_type = 'income_statement'
            elif 'cash_flow' in filename.lower():
                statement_type = 'cash_flow'
            else:
                continue
                
            file_path = os.path.join(self.statements_dir, filename)
            self.files[statement_type][period_type] = file_path
            
        for statement_type, periods in self.files.items():
            print(f"{statement_type}: {periods}")
        print(f"Company ticker: {self.company_ticker}\n")

    def _read_excel(self, file_path):
        """Read Excel file with proper handling of headers."""
        if file_path is None:
            return None
            
        if not os.path.exists(file_path):
            print(f"Warning: File does not exist: {file_path}")
            return None
            
        try:
            # Try reading the Excel file
            df = pd.read_excel(file_path)
            
            # Check if we have unnamed columns
            unnamed_cols = [col for col in df.columns if 'Unnamed' in str(col)]
            
            # If we have unnamed columns, try reading with header on row 2
            if len(unnamed_cols) > 0:
                print(f"Detected unnamed columns in {os.path.basename(file_path)}, trying with header row 2")
                df = pd.read_excel(file_path, header=1)
            
            # Debug info
            print(f"Successfully read {os.path.basename(file_path)}, shape: {df.shape}")
            print(f"Columns: {df.columns.tolist()}")
            
            return df
        except Exception as e:
            print(f"Error reading file {file_path}: {str(e)}")
            return None

    def _extract_years_from_cols(self, df):
        """Extract year information from dataframe columns."""
        years = []
        for col in df.columns[1:]:  # Skip the first column (account names)
            # Try to find a year in the column name
            year_match = re.search(r'20\d{2}', str(col))
            if year_match:
                years.append(year_match.group(0))
            else:
                # If no year found, just use a placeholder
                years.append('Unknown')
        
        return years

    def consolidate_statements(self):
        """Create a single consolidated dataframe with all statements."""
        # Step 1: Read all files
        data = {}
        for statement_type in ['balance_sheet', 'income_statement', 'cash_flow']:
            data[statement_type] = {
                'QTR': self._read_excel(self.files[statement_type]['QTR']),
                'FY': self._read_excel(self.files[statement_type]['FY'])
            }
        
        # Extract years from the data for reference
        all_years = set()
        for statement_type in data:
            for period_type in data[statement_type]:
                if data[statement_type][period_type] is not None:
                    years = self._extract_years_from_cols(data[statement_type][period_type])
                    all_years.update(years)
        
        all_years = sorted(list(all_years))
        print(f"Found years: {all_years}")
        
        # Placeholder for the consolidated dataframe
        consolidated_df = pd.DataFrame()
        
        # Placeholder for section ranges - will be needed for styling
        self.section_ranges = {}  # Changed from local variable to instance attribute
        
        # List of quarterly dataframes to process
        quarterly_dataframes = [data['balance_sheet']['QTR'], data['income_statement']['QTR'], data['cash_flow']['QTR']]
        
        # TODO: Fill consolidated_df and section_ranges
        for i, df in enumerate(quarterly_dataframes):
            if df is None:
                continue
            
            # Create a dictionary to track where each quarter should be inserted
            account_col = df.columns[0]  # First column is account names
            data_cols = list(df.columns[1:])  # All other columns are data
            
            for year in all_years:
                # Create a list of quarters we need for this year
                quarters = ['Q1', 'Q2', 'Q3', 'Q4']
                
                # Check which quarters already exist
                existing_qtrs = [q for q in quarters if f"{q} {year}" in data_cols]
                
                # Find quarters that need to be added
                for quarter in quarters:
                    qtr_col = f"{quarter} {year}"
                    if qtr_col not in data_cols:
                        # Determine where to insert it
                        qtr_num = int(quarter[1])  # Extract the quarter number
                        
                        # Find position based on order of quarters (1, 2, 3, 4)
                        insert_pos = 1  # Start after the account column
                        for existing_col in data_cols:
                            # Check if the existing column is for a year/quarter before our target
                            if existing_col.endswith(year):
                                existing_qtr = existing_col.split()[0]
                                if existing_qtr in quarters:
                                    existing_qtr_num = int(existing_qtr[1])
                                    if existing_qtr_num < qtr_num:
                                        # This existing quarter comes before our target
                                        insert_pos = data_cols.index(existing_col) + 2  # +1 for index, +1 for account col
                        
                        # Insert the missing quarter column at the calculated position
                        print(f"Inserting {qtr_col} at position {insert_pos}")
                        df.insert(loc=insert_pos, column=qtr_col, value=None)
                        
                        # Update data_cols to reflect the new column
                        data_cols.insert(insert_pos-1, qtr_col)

        # For each row in the quarterly balance sheet
        for row_idx, row in data['balance_sheet']['QTR'].iterrows():
            account_name = row[data['balance_sheet']['QTR'].columns[0]]  # Get account name
            
            # Look for this account in the yearly data
            if data['balance_sheet']['FY'] is not None:
                yearly_df = data['balance_sheet']['FY']
                yearly_col = yearly_df.columns[0]  # Account column name
                yearly_matches = yearly_df[yearly_df[yearly_col] == account_name]
                
                if not yearly_matches.empty:
                    # Get the matching row from yearly data
                    yearly_row = yearly_matches.iloc[0]
                    
                    # For each year column in yearly data
                    for i, col in enumerate(yearly_df.columns[1:], 1):
                        if 'FY' in col:
                            year = re.search(r'20\d{2}', col).group(0) if re.search(r'20\d{2}', col) else None
                            if year:
                                q4_col = f"Q4 {year}"
                                
                                # If Q4 column exists in quarterly data
                                if q4_col in data['balance_sheet']['QTR'].columns:
                                    # Copy the yearly value to Q4
                                    data['balance_sheet']['QTR'].at[row_idx, q4_col] = yearly_row.iloc[i]

        # Now simply concatenate all dataframes to create the consolidated dataframe
        row_counter = 0
        dfs_to_concat = []
        modified_dfs = {}  # Store references to the modified dataframes
        
        # Make sure all dataframes have consistent column names first
        for statement_type in ['balance_sheet', 'income_statement', 'cash_flow']:
            if data[statement_type]['QTR'] is not None:
                # Rename first column to 'Account' for consistency
                first_col = data[statement_type]['QTR'].columns[0]
                data[statement_type]['QTR'].rename(columns={first_col: 'Account'}, inplace=True)
                modified_dfs[statement_type] = data[statement_type]['QTR']
        
        # Track section ranges for styling
        for statement_type in ['balance_sheet', 'income_statement', 'cash_flow']:
            df = modified_dfs.get(statement_type)
            if df is not None:
                # Record section range
                start_row = row_counter
                row_counter += len(df)
                end_row = row_counter - 1
                self.section_ranges[statement_type] = (start_row, end_row)
                dfs_to_concat.append(df)
                print(f"Added {statement_type} rows {start_row}-{end_row}")
        
        # Concatenate all dataframes
        if dfs_to_concat:
            consolidated_df = pd.concat(dfs_to_concat, ignore_index=True)
            
            # Sort quarterly columns chronologically
            account_col = consolidated_df.columns[0]  # First column (Account names)
            other_cols = consolidated_df.columns[1:]
            
            # Extract quarter and year information for sorting
            quarter_info = []
            non_quarter_cols = []
            
            for col in other_cols:
                match = re.search(r'(Q\d)\s+(\d{4})', str(col))
                if match:
                    quarter = match.group(1)
                    year = match.group(2)
                    quarter_num = int(quarter[1])
                    quarter_info.append((int(year), quarter_num, col))
                else:
                    non_quarter_cols.append(col)
            
            # Sort quarterly columns chronologically
            quarter_info.sort()
            
            # Reconstruct column list in the right order
            new_cols = [account_col]
            for _, _, col in quarter_info:
                new_cols.append(col)
            
            # Add any non-quarterly columns
            new_cols.extend(non_quarter_cols)
            
            # Reorder the dataframe columns
            consolidated_df = consolidated_df[new_cols]
            print(f"Sorted quarterly columns chronologically: {new_cols}")
            
            print(f"Created consolidated dataframe with shape: {consolidated_df.shape}")
        else:
            print("No dataframes to concatenate")
            return None
        
        # Add yearly data to the consolidated dataframe
        consolidated_df = self._add_yearly_data(consolidated_df)
        
        # Add calculated metrics columns
        consolidated_df = self._add_calculated_columns(consolidated_df)
        
        # Save the consolidated dataframe
        self._save_consolidated_workbook(consolidated_df, self.section_ranges)
        
        return consolidated_df

    def _save_consolidated_workbook(self, df, section_ranges):
        """Save the consolidated dataframe to an Excel workbook."""
        if df is None or df.empty:
            print("No data to save")
            return
        
        # Debug print section ranges    
        print("Section ranges for styling:")
        for section, (start, end) in section_ranges.items():
            print(f"  {section}: rows {start}-{end}")
            
        output_path = os.path.join(os.path.dirname(self.files['balance_sheet']['FY']), 
                               f"consolidated_statements_{self.company_ticker}.xlsx" if self.company_ticker else "consolidated_statements.xlsx")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidated Statements"
        
        # Define styles
        header_style = Font(bold=True)
        section_colors = {
            'balance_sheet': "C6EFCE",  # Light green
            'income_statement': "FFEB9C",  # Light yellow
            'cash_flow': "FFC7CE"  # Light red
        }
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Start writing data
        current_row = 1
        
        # Write title in first column only
        title = f"Consolidated Financial Statements - {self.company_ticker}" if self.company_ticker else "Consolidated Financial Statements"
        ws.cell(row=current_row, column=1, value=title)
        
        # No merging cells - apply styling only to the first cell
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="left")
        
        current_row += 2  # Add space after title
        
        # Write column headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = header_style
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        current_row += 1
        
        # Pre-determine which rows belong to which section
        row_section_map = {}
        for row_idx in range(len(df)):
            for section, (start, end) in section_ranges.items():
                if start <= row_idx <= end:
                    row_section_map[row_idx] = section
                    break
        
        # Track when we add section headers
        section_headers_added = set()
        
        # Write rows with section styling
        for row_idx, row in df.iterrows():
            # Get section type from our pre-computed map
            section_type = row_section_map.get(row_idx)
            
            # If this is the first row of a section, add a section header
            if section_type and section_type not in section_headers_added:
                # Insert a section header
                header_row = current_row
                section_name = section_type.replace('_', ' ').title()
                ws.cell(row=header_row, column=1, value=section_name)
                
                # Apply formatting to just the first cell (no cell merging)
                header_cell = ws.cell(row=header_row, column=1)
                header_cell.font = Font(bold=True)
                header_cell.fill = PatternFill(start_color=section_colors[section_type], 
                                              end_color=section_colors[section_type], 
                                              fill_type="solid")
                header_cell.alignment = Alignment(horizontal="left")
                header_cell.border = thin_border
                
                # Add empty cells for the rest of the row
                for col_idx in range(2, len(df.columns) + 1):
                    empty_cell = ws.cell(row=header_row, column=col_idx, value="")
                    empty_cell.border = thin_border
                
                current_row += 1
                section_headers_added.add(section_type)
                print(f"Added section header for {section_type} at Excel row {header_row}")
            
            # Write data row
            for col_idx, col_name in enumerate(df.columns, 1):
                value = row[col_name]
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = thin_border
                
                # Format numbers
                if col_idx > 1 and value is not None:  # Skip account name column
                    try:
                        float_val = float(value)
                        cell.number_format = '#,##0.00'
                    except (ValueError, TypeError):
                        pass
            
            current_row += 1
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(output_path)
        print(f"Consolidated statements saved to {output_path}")

    def _add_yearly_data(self, df):
        """Add in yearly data to the consolidated dataframe."""
        if df is None or df.empty:
            print("No consolidated dataframe to add yearly data to")
            return df
        
        print("Adding yearly financial data to consolidated dataframe...")
        
        # First, identify all quarterly columns and ensure they're sorted chronologically
        quarter_columns = []
        quarter_info = []
        
        for col in df.columns:
            # Match quarterly columns like "Q1 2023"
            match = re.search(r'(Q\d)\s+(\d{4})', str(col))
            if match:
                quarter = match.group(1)
                year = match.group(2)
                quarter_num = int(quarter[1])  # Extract just the number
                # Create tuple for sorting: (year, quarter_number, original_column_name)
                quarter_info.append((int(year), quarter_num, col))
        
        # Sort quarterly columns chronologically
        quarter_info.sort()  # This sorts by year first, then by quarter number
        
        # Create new DataFrame with properly ordered columns
        account_col = df.columns[0]  # Usually 'Account'
        new_cols = [account_col]
        
        # Add all quarterly columns in correct order
        for _, _, col_name in quarter_info:
            new_cols.append(col_name)
        
        # Add any remaining non-quarterly columns
        for col in df.columns:
            if col not in new_cols:
                new_cols.append(col)
        
        # Create new DataFrame with reordered columns
        df = df[new_cols]
        print(f"Reordered columns: {new_cols}")
        
        # Process each statement type
        for statement_type in ['balance_sheet', 'income_statement', 'cash_flow']:
            # Check if we have yearly data for this statement type
            if self.files[statement_type]['FY'] is None:
                print(f"No yearly data for {statement_type}, skipping")
                continue
            
            # Read yearly data
            yearly_df = self._read_excel(self.files[statement_type]['FY'])
            if yearly_df is None:
                print(f"Could not read yearly data for {statement_type}, skipping")
                continue
            
            # Rename first column to 'Account' for consistency
            first_col = yearly_df.columns[0]
            yearly_df.rename(columns={first_col: 'Account'}, inplace=True)
            
            # Process FY columns to ensure they're valid
            fy_columns = []
            for col in yearly_df.columns[1:]:  # Skip 'Account' column
                # Validate this is a proper FY column with a year
                year_match = re.search(r'FY\s?(\d{4})', str(col))
                if year_match:
                    year = year_match.group(1)
                    # Standardize the column name format
                    std_col_name = f"FY {year}"
                    fy_columns.append((std_col_name, year, col))
                    print(f"Found yearly column: {col} → {std_col_name}")
                else:
                    print(f"Skipping non-standard column: {col}")
            
            # Sort FY columns by year
            fy_columns.sort(key=lambda x: x[1])
            
            # Find the last quarterly column - we'll insert all FY columns after this
            last_qtr_idx = 0
            for i, col in enumerate(df.columns):
                if col in quarter_columns:
                    last_qtr_idx = i
            
            # Add all FY columns after the last quarterly column, in order of year
            for idx, (std_col_name, year, orig_col) in enumerate(fy_columns):
                if std_col_name not in df.columns:
                    # Insert position is after all quarterly columns plus any FY columns we've already added
                    insert_pos = last_qtr_idx + 1 + idx
                    
                    print(f"Adding column {std_col_name} to consolidated dataframe at position {insert_pos}")
                    # Add column to DataFrame at the specified position
                    df.insert(insert_pos, std_col_name, None)
            
            # For each row in yearly_df, try to find a matching row in df
            for _, yearly_row in yearly_df.iterrows():
                account_name = yearly_row['Account']
                
                # Find matching rows in consolidated df
                matching_rows = df[df['Account'] == account_name]
                
                if not matching_rows.empty:
                    # If account exists, update the FY columns
                    row_idx = matching_rows.index[0]
                    for std_col_name, year, orig_col in fy_columns:
                        if std_col_name in df.columns:
                            df.at[row_idx, std_col_name] = yearly_row[orig_col]
                else:
                    # If account doesn't exist, we need to add it to the appropriate section
                    # First, determine which section this belongs to
                    section_row = pd.Series([None] * len(df.columns), index=df.columns)
                    section_row['Account'] = account_name
                    
                    # Copy data from yearly row to new row
                    for std_col_name, year, orig_col in fy_columns:
                        if std_col_name in df.columns:
                            section_row[std_col_name] = yearly_row[orig_col]
                    
                    # Find the end of the appropriate section to insert the row
                    if hasattr(self, 'section_ranges') and statement_type in self.section_ranges:
                        insert_idx = self.section_ranges[statement_type][1] + 1  # End of section + 1
                        print(f"Inserting account {account_name} at position {insert_idx}")
                        # Insert the new row
                        df = pd.concat([df.iloc[:insert_idx], 
                                       pd.DataFrame([section_row]), 
                                       df.iloc[insert_idx:]]).reset_index(drop=True)
                        
                        # Update section ranges to reflect the insertion
                        for section in self.section_ranges:
                            if section == statement_type:
                                self.section_ranges[section] = (
                                    self.section_ranges[section][0], 
                                    self.section_ranges[section][1] + 1
                                )
                            elif self.section_ranges[section][0] > insert_idx:
                                self.section_ranges[section] = (
                                    self.section_ranges[section][0] + 1, 
                                    self.section_ranges[section][1] + 1
                                )
                    else:
                        # If we don't have section ranges, just append to the end
                        print(f"Appending account {account_name} to the end")
                        # Use concat instead of append (deprecated)
                        df = pd.concat([df, pd.DataFrame([section_row])], ignore_index=True)
        
        print("Yearly data addition complete")
        return df

    def _add_calculated_columns(self, df):
        """Add in calculated columns to the consolidated dataframe."""
        if df is None or df.empty:
            print("No data to add calculated columns to")
            return df
        
        print("Adding calculated financial metrics...")
        
        # Create a copy of the dataframe to avoid fragmentation
        df = df.copy()
        
        # Debug: Print some sample account names to help identify the correct patterns
        print("Sample account names in the dataframe:")
        for account in list(df['Account'].unique())[:10]:  # Show first 10 accounts
            print(f"  - {account}")
        
        # Get all quarter/year columns (Q1 2023, Q2 2023, etc.)
        data_cols = [col for col in df.columns if re.search(r'(Q\d|FY)\s+\d{4}', str(col))]
        print(f"Found {len(data_cols)} data columns: {data_cols}")
        
        # More comprehensive patterns for account matching, based on actual data
        account_patterns = {
            'current_assets': ['current assets', 'total current assets'],
            'current_liabilities': ['current liabilities', 'total current liabilities'],
            'revenue': ['revenue', 'total revenue', 'net sales', 'current deferred revenue'],
            'capex': ['Purchase of PP&E','capital expenditure', 'capex', 'cap ex', 'purchases of property', 
                     'additions to property and equipment', 'payments for property and equipment'],
            'op_cash_flow': ['operating cash flow', 'Net Cash from Operations', 'Net Cash from Continuing Operating Activities','cash from operating', 
                             'operating activities', 'cash provided by operating', 
                             'net cash from continuing operating activities']
        }
        
        # Find the best match for each account type
        account_matches = {}
        for account_type, patterns in account_patterns.items():
            for pattern in patterns:
                matches = df[df['Account'].str.contains(pattern, case=False, na=False)]
                if not matches.empty:
                    # Store the first match
                    account_matches[account_type] = matches['Account'].iloc[0]
                    print(f"Found match for {account_type}: '{account_matches[account_type]}'")
                    break
        
        if not account_matches:
            print(f"WARNING: No match found for any account type")
        
        # Add new rows to the dataframe for calculated metrics
        new_rows = []
        
        # Create the Working Capital row
        wc_row = pd.Series({'Account': 'Working Capital'})
        new_rows.append(wc_row)
        
        # Create Working Capital % of Revenue row
        wc_pct_row = pd.Series({'Account': 'Working Capital % of Revenue'})
        new_rows.append(wc_pct_row)
        
        # Create CapEx % of Revenue row
        capex_pct_row = pd.Series({'Account': 'CapEx % of Revenue'})
        new_rows.append(capex_pct_row)
        
        # Create Free Cash Flow row
        fcf_row = pd.Series({'Account': 'Free Cash Flow'})
        new_rows.append(fcf_row)
        
        # Add these rows to the dataframe
        new_df_rows = pd.DataFrame(new_rows)
        df = pd.concat([df, new_df_rows], ignore_index=True)
        
        # Process each data column and calculate metrics
        for col in data_cols:
            print(f"Processing metrics for {col}...")
            
            # 1. Working Capital
            if 'current_assets' in account_matches and 'current_liabilities' in account_matches:
                try:
                    # Get the actual row values
                    assets_row = df[df['Account'] == account_matches['current_assets']]
                    liab_row = df[df['Account'] == account_matches['current_liabilities']]
                    
                    if not assets_row.empty and not liab_row.empty and pd.notna(assets_row[col].iloc[0]) and pd.notna(liab_row[col].iloc[0]):
                        current_assets = float(assets_row[col].iloc[0])
                        current_liabilities = float(liab_row[col].iloc[0])
                        
                        # Calculate working capital
                        working_capital = current_assets - current_liabilities
                        
                        # Add to the Working Capital row directly
                        wc_idx = df[df['Account'] == 'Working Capital'].index[0]
                        df.loc[wc_idx, col] = working_capital
                        
                        print(f"  Working Capital for {col}: {working_capital:,.2f}")
                        
                        # 2. Working Capital as % of Revenue
                        if 'revenue' in account_matches:
                            rev_row = df[df['Account'] == account_matches['revenue']]
                            
                            if not rev_row.empty and pd.notna(rev_row[col].iloc[0]):
                                revenue = float(rev_row[col].iloc[0])
                                
                                if revenue != 0:  # Avoid division by zero
                                    wc_percent = (working_capital / revenue) * 100
                                    
                                    # Add to the Working Capital % row directly
                                    wc_pct_idx = df[df['Account'] == 'Working Capital % of Revenue'].index[0]
                                    df.loc[wc_pct_idx, col] = wc_percent
                                    
                                    print(f"  Working Capital % for {col}: {wc_percent:.2f}%")
                except Exception as e:
                    print(f"  Error calculating Working Capital for {col}: {str(e)}")
            
            # 3. CapEx % of Revenue
            if 'capex' in account_matches and 'revenue' in account_matches:
                try:
                    capex_row = df[df['Account'] == account_matches['capex']]
                    rev_row = df[df['Account'] == account_matches['revenue']]
                    
                    if not capex_row.empty and not rev_row.empty and pd.notna(capex_row[col].iloc[0]) and pd.notna(rev_row[col].iloc[0]):
                        capex = float(capex_row[col].iloc[0])
                        revenue = float(rev_row[col].iloc[0])
                        
                        capex_abs = abs(capex)  # Use absolute value for percentage
                        
                        if revenue != 0:  # Avoid division by zero
                            capex_percent = (capex_abs / revenue) * 100
                            
                            # Add to the CapEx % row directly
                            capex_pct_idx = df[df['Account'] == 'CapEx % of Revenue'].index[0]
                            df.loc[capex_pct_idx, col] = capex_percent
                            
                            print(f"  CapEx % for {col}: {capex_percent:.2f}%")
                except Exception as e:
                    print(f"  Error calculating CapEx % for {col}: {str(e)}")
            
            # 4. Free Cash Flow
            if 'op_cash_flow' in account_matches and 'capex' in account_matches:
                try:
                    ocf_row = df[df['Account'] == account_matches['op_cash_flow']]
                    capex_row = df[df['Account'] == account_matches['capex']]
                    
                    if not ocf_row.empty and not capex_row.empty and pd.notna(ocf_row[col].iloc[0]) and pd.notna(capex_row[col].iloc[0]):
                        op_cash_flow = float(ocf_row[col].iloc[0])
                        capex = float(capex_row[col].iloc[0])
                        
                        # Calculate FCF (CapEx is often negative in statements)
                        fcf = op_cash_flow + capex if capex < 0 else op_cash_flow - capex
                        
                        # Add to the Free Cash Flow row directly
                        fcf_idx = df[df['Account'] == 'Free Cash Flow'].index[0]
                        df.loc[fcf_idx, col] = fcf
                        
                        print(f"  Free Cash Flow for {col}: {fcf:,.2f}")
                except Exception as e:
                    print(f"  Error calculating Free Cash Flow for {col}: {str(e)}")
        
        print("Calculated metrics added successfully")
        return df

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Consolidate financial statements for a specific ticker.')
    parser.add_argument('ticker', nargs='?', help='Stock ticker symbol')
    args = parser.parse_args()
    
    # Get the ticker either from command line or user input
    ticker = args.ticker
    if not ticker:
        ticker = input("Please enter the ticker symbol: ").strip().upper()
        if not ticker:
            print("No ticker provided. Exiting.")
            sys.exit(1)
    else:
        ticker = ticker.upper()
    
    print(f"Processing financial statements for ticker: {ticker}")
    
    # Construct the statements directory path with the ticker subfolder
    script_dir = os.path.dirname(os.path.abspath(__file__))
    statements_dir = os.path.join(script_dir, 'statements', ticker)
    
    # Check if the directory exists
    if not os.path.exists(statements_dir):
        print(f"Error: No directory found for ticker {ticker} at path: {statements_dir}")
        print("Available tickers:")
        tickers_dir = os.path.join(script_dir, 'statements')
        available_tickers = [d for d in os.listdir(tickers_dir) if os.path.isdir(os.path.join(tickers_dir, d))]
        for t in available_tickers:
            print(f"  - {t}")
        sys.exit(1)
    
    # Initialize consolidator with the ticker-specific directory
    consolidator = FinancialStatementConsolidator(statements_dir)
    consolidator.consolidate_statements()
    
    print("Financial statement consolidation complete!")

if __name__ == "__main__":
    main()